# ═══════════════════════════════════════
# Parsers — Excel Parser
# ═══════════════════════════════════════
import openpyxl
from typing import Dict, List, Optional
from io import BytesIO
from core.logger import logger
from core.exceptions import ValidationError


class ExcelTranscriptParser:
    """
    Parse Excel transcript files (.xlsx).
    Supports both bulk tabular student lists and individual student transcript reports.
    """

    REQUIRED_HEADERS = {"رقم الطالب", "اسم الطالب", "كود المادة", "اسم المادة", "التقدير"}

    def parse(self, file_data: bytes) -> List[Dict]:
        """Parse Excel file and return list of student records."""
        try:
            # Load with data_only=True to get evaluated values of formulas
            wb = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
            ws = wb.active
            if not ws:
                raise ValidationError("ملف Excel فارغ")

            # Try to read headers in first few rows to locate tabular format
            is_tabular = False
            header_row_idx = 0
            header_map = {}
            for r in range(1, 6):
                row_vals = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[r]]
                missing = self.REQUIRED_HEADERS - set(row_vals)
                if not missing:
                    is_tabular = True
                    header_row_idx = r
                    header_map = {h: i for i, h in enumerate(row_vals)}
                    break

            if is_tabular:
                logger.info("Excel parser: Detected bulk tabular format")
                students = {}
                for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                    if len(row) <= max(header_map.values()):
                        continue
                    student_code = str(row[header_map.get("رقم الطالب", 0)] or "").strip()
                    if not student_code:
                        continue

                    if student_code not in students:
                        students[student_code] = {
                            "student_code": student_code,
                            "name": str(row[header_map.get("اسم الطالب", 1)] or "").strip(),
                            "courses": [],
                        }

                    students[student_code]["courses"].append({
                        "code": str(row[header_map.get("كود المادة", 2)] or "").strip(),
                        "name": str(row[header_map.get("اسم المادة", 3)] or "").strip(),
                        "grade": str(row[header_map.get("التقدير", 4)] or "").strip(),
                        "credit_hours": int(row[header_map.get("الساعات", 5)] or 0) if "الساعات" in header_map and header_map["الساعات"] < len(row) else 0,
                    })

                wb.close()
                logger.info(f"Parsed {len(students)} students from bulk Excel")
                return list(students.values())

            # If not tabular, search for individual transcript format
            student_code = None
            student_name = None
            
            # Scan first 30 rows for student info
            max_scan = min(30, ws.max_row + 1) if ws.max_row else 30
            for r in range(1, max_scan):
                row_vals = [cell.value for cell in ws[r]]
                for val in row_vals:
                    if val and isinstance(val, str):
                        if "كود الطالب" in val:
                            student_code = val.split(":")[-1].strip()
                        if "أسم الطالب" in val or "اسم الطالب" in val:
                            student_name = val.split(":")[-1].strip()

            if student_code:
                logger.info(f"Excel parser: Detected individual transcript format for student {student_code}")
                courses = []
                r = 1
                max_row = ws.max_row if ws.max_row else 1000
                while r <= max_row:
                    row_vals = [cell.value for cell in ws[r]]
                    row_strs = [str(v).strip() if v is not None else "" for v in row_vals]
                    
                    # Check if this row is a course header row
                    if any("اسم المقرر" in s or "اسم المادة" in s for s in row_strs) and any("كود المقرر" in s or "كود المادة" in s for s in row_strs):
                        # Locate indices
                        header_map = {}
                        for idx, val in enumerate(row_vals):
                            if val is None:
                                continue
                            val_str = str(val).strip()
                            if "كود المقرر" in val_str or "كود المادة" in val_str:
                                header_map["code"] = idx
                            elif "اسم المقرر" in val_str or "اسم المادة" in val_str:
                                header_map["name"] = idx
                            elif "تقدير" in val_str or "التقدير" in val_str:
                                header_map["grade"] = idx
                            elif "ساعات" in val_str or "الساعات" in val_str:
                                header_map["hours"] = idx
                        
                        r += 1
                        while r <= max_row:
                            course_row = [cell.value for cell in ws[r]]
                            if not any(course_row):
                                r += 1
                                continue
                            
                            if "name" not in header_map or header_map["name"] >= len(course_row):
                                break
                            name_val = course_row[header_map["name"]]
                            if name_val is None:
                                break
                            
                            name_str = str(name_val).strip()
                            if "المجمـــوع" in name_str or "المجموع" in name_str or name_str == "":
                                break
                                
                            code_val = course_row[header_map["code"]] if "code" in header_map and header_map["code"] < len(course_row) else ""
                            grade_val = course_row[header_map["grade"]] if "grade" in header_map and header_map["grade"] < len(course_row) else ""
                            hours_val = course_row[header_map["hours"]] if "hours" in header_map and header_map["hours"] < len(course_row) else 0
                            
                            try:
                                hours = int(float(hours_val)) if hours_val else 0
                            except ValueError:
                                hours = 0
                                
                            courses.append({
                                "code": str(code_val).strip(),
                                "name": name_str,
                                "grade": str(grade_val).strip(),
                                "credit_hours": hours
                            })
                            r += 1
                    else:
                        r += 1
                
                wb.close()
                logger.info(f"Parsed individual transcript with {len(courses)} courses")
                return [{
                    "student_code": student_code,
                    "name": student_name or "طالب غير معروف",
                    "courses": courses
                }]

            wb.close()
            raise ValidationError(
                "صيغة ملف Excel غير مدعومة. يرجى التأكد من وجود الأعمدة المطلوبة "
                "(رقم الطالب، اسم الطالب، كود المادة، اسم المادة، التقدير) "
                "أو أن يكون الملف تقريراً أكاديمياً يحتوي على كود الطالب وأسم الطالب."
            )

        except ValidationError:
            raise
        except Exception as e:
            logger.error(f"Excel parsing error: {e}")
            raise ValidationError(f"خطأ في قراءة ملف Excel: {str(e)}")
