import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Tuple
from datetime import datetime
import uuid
import logging
import re

logger = logging.getLogger("acadexa.parser")

def clean(value):
    if value is None:
        return ""
    return str(value).strip()

def get_cell(ws, col_letter, row_idx):
    try:
        return ws[f"{col_letter}{row_idx}"].value
    except:
        return None

def find_first_semester_row(ws):
    """Find the first semester data row"""
    for row in range(1, min(100, ws.max_row + 1)):
        for col in range(1, min(ws.max_column + 1, 30)):
            val = clean(ws.cell(row, col).value)
            if 'القسم/الشعبة' in val:
                return row, col
    return None, None

def detect_offset(ws):
    """
    Determine the offset between the current sheet and the reference sheet
    based on the location of the first cell containing 'القسم/الشعبة'
    """
    ref_row, ref_col = find_first_semester_row(ws)
    if ref_row and ref_col:
        row_offset = ref_row - 40
        col_offset = ref_col - 11
        return row_offset, col_offset
    return 0, 0

def get_cell_with_offset(ws, base_col_letter, base_row, row_offset, col_offset):
    """Get cell with offset applied"""
    base_col_num = openpyxl.utils.column_index_from_string(base_col_letter)
    new_col_num = base_col_num + col_offset
    new_col_letter = openpyxl.utils.get_column_letter(new_col_num)
    new_row = base_row + row_offset
    return get_cell(ws, new_col_letter, new_row)

def parse_student_info_old(ws):
    """Parse student information from first rows"""
    student = {
        "id": "",
        "name": "",
        "study_level": "",
        "cumulative_percentage": "",
    }
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            v = clean(cell.value)
            if v.startswith("كود الطالب :"):
                student["id"] = v.replace("كود الطالب :", "").strip()
            elif v.startswith("أسم الطالب :"):
                student["name"] = v.replace("أسم الطالب :", "").strip()
            elif v.startswith("مستوى الدراسة :"):
                student["study_level"] = v.replace("مستوى الدراسة :", "").strip()
            elif v.startswith("النسبة(بحساب النقاط) :"):
                student["cumulative_percentage"] = v.replace("النسبة(بحساب النقاط) :", "").strip()
    return student

def parse_semesters_with_offset(ws):
    """Parse semesters with automatic offset support"""
    
    row_offset, col_offset = detect_offset(ws)
    
    semesters = []
    current_semester = None
    current_courses = []
    
    COLS = {
        'K': 'K', 'AH': 'AH', 'BO': 'BO',
        'I': 'I', 'Y': 'Y', 'M': 'M', 'X': 'X',
        'L': 'L', 'W': 'W', 'AN': 'AN',
        'CD': 'CD', 'BC': 'BC', 'BU': 'BU',
        'AL': 'AL', 'Z': 'Z', 'AC': 'AC',
        'AS': 'AS', 'AU': 'AU', 'O': 'O', 'Q': 'Q'
    }
    
    start_row = max(1, 38 + row_offset)
    
    for row_idx in range(start_row, ws.max_row + 1):
        cell_k = get_cell_with_offset(ws, COLS['K'], row_idx, row_offset, col_offset)
        cell_ah = get_cell_with_offset(ws, COLS['AH'], row_idx, row_offset, col_offset)
        cell_bo = get_cell_with_offset(ws, COLS['BO'], row_idx, row_offset, col_offset)
        
        cell_k_clean = clean(cell_k)
        
        if 'القسم/الشعبة :' in cell_k_clean:
            if current_semester is not None:
                current_semester['courses'] = current_courses
                semesters.append(current_semester)
            
            current_semester = {
                "department": cell_k_clean.replace('القسم/الشعبة :', '').strip(),
                "level_semester": clean(cell_ah).replace('المستوى/الفصل :', '').strip(),
                "academic_year": clean(cell_bo).replace('العام الأكاديمي   :', '').strip(),
                "total_passed_hours": "",
                "gpa": "",
                "grade": "",
                "semester_hours": "",
                "semester_gpa": "",
                "level_hours": "",
                "level_gpa": "",
            }
            current_courses = []
            continue
        
        if current_semester is None:
            continue
        
        cell_i = clean(get_cell_with_offset(ws, COLS['I'], row_idx, row_offset, col_offset))
        cell_y = clean(get_cell_with_offset(ws, COLS['Y'], row_idx, row_offset, col_offset))
        cell_an = clean(get_cell_with_offset(ws, COLS['AN'], row_idx, row_offset, col_offset))
        
        if cell_i.startswith('الساعات المجتازة :') and cell_y.startswith('النقاط :'):
            current_semester['total_passed_hours'] = cell_i.replace('الساعات المجتازة :', '').strip()
            current_semester['gpa'] = cell_y.replace('النقاط :', '').strip()
            current_semester['grade'] = cell_an.replace('التقدير :', '').strip()
        
        cell_m = clean(get_cell_with_offset(ws, COLS['M'], row_idx, row_offset, col_offset))
        cell_x = clean(get_cell_with_offset(ws, COLS['X'], row_idx, row_offset, col_offset))
        
        if cell_m.startswith('الساعات المجتازة :') and cell_x.startswith('النقاط :'):
            current_semester['semester_hours'] = cell_m.replace('الساعات المجتازة :', '').strip()
            current_semester['semester_gpa'] = cell_x.replace('النقاط :', '').strip()
        
        cell_l = clean(get_cell_with_offset(ws, COLS['L'], row_idx, row_offset, col_offset))
        cell_w = clean(get_cell_with_offset(ws, COLS['W'], row_idx, row_offset, col_offset))
        
        if cell_l.startswith('الساعات المجتازة :') and cell_w.startswith('النقاط :'):
            current_semester['level_hours'] = cell_l.replace('الساعات المجتازة :', '').strip()
            current_semester['level_gpa'] = cell_w.replace('النقاط :', '').strip()
        
        cell_cd = clean(get_cell_with_offset(ws, COLS['CD'], row_idx, row_offset, col_offset))
        cell_bc = clean(get_cell_with_offset(ws, COLS['BC'], row_idx, row_offset, col_offset))
        
        if cell_cd == 'م' or cell_bc == 'المجمـــوع' or cell_bc == 'المجموع':
            continue
        
        if cell_cd.isdigit() and cell_bc:
            course = {
                "seq": cell_cd,
                "course_code": clean(get_cell_with_offset(ws, COLS['BU'], row_idx, row_offset, col_offset)),
                "course_name": cell_bc,
                "passed": clean(get_cell_with_offset(ws, COLS['I'], row_idx, row_offset, col_offset)),
                "grade_letter": clean(get_cell_with_offset(ws, COLS['O'], row_idx, row_offset, col_offset)),
                "score": clean(get_cell_with_offset(ws, COLS['Q'], row_idx, row_offset, col_offset)),
                "hours": clean(get_cell_with_offset(ws, COLS['AL'], row_idx, row_offset, col_offset)),
                "points": clean(get_cell_with_offset(ws, COLS['Z'], row_idx, row_offset, col_offset)),
                "cumulative": clean(get_cell_with_offset(ws, COLS['AC'], row_idx, row_offset, col_offset)),
                "min_score": clean(get_cell_with_offset(ws, COLS['AS'], row_idx, row_offset, col_offset)),
                "max_score": clean(get_cell_with_offset(ws, COLS['AU'], row_idx, row_offset, col_offset)),
            }
            current_courses.append(course)
    
    if current_semester is not None:
        current_semester['courses'] = current_courses
        semesters.append(current_semester)
    
    return semesters


class ExcelParser:
    """Excel parser that reads all sheets and saves directly to Supabase"""
    
    def __init__(self, file_path: Path, department_code: str = ""):
        self.file_path = file_path
        self.department_code = department_code
        self.stats = {"students": 0, "courses": 0, "semesters": 0}
        self.errors = []
        self._supabase = None
        self._department_id = None  # <-- جديد: تخزين الـ department_id مؤقتاً
        
    def _get_supabase(self):
        """Lazy load Supabase client"""
        if self._supabase is None:
            try:
                from core.db.supabase_client import supabase
                self._supabase = supabase
            except ImportError:
                logger.warning("Supabase client not available, using JSON output only")
                self._supabase = None
        return self._supabase
    
    def _get_department_id(self):
        """Get department ID from department code (cached)"""
        if self._department_id is not None:
            return self._department_id
        
        if not self.department_code:
            return None
        
        try:
            supabase = self._get_supabase()
            if supabase:
                result = supabase.table("departments").select("id").eq("code", self.department_code).execute()
                if result.data:
                    self._department_id = result.data[0]["id"]
                    logger.info(f"Found department ID {self._department_id} for code {self.department_code}")
                    return self._department_id
        except Exception as e:
            logger.warning(f"Could not find department for code {self.department_code}: {e}")
        
        return None
    
    def parse_all_students(self) -> Tuple[List[Dict], List[Dict]]:
        """Parse all students and save to Supabase if available"""
        students = []
        errors = []
        
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            logger.info(f"Opened workbook with {len(wb.sheetnames)} sheets")
            
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    
                    student_info = parse_student_info_old(ws)
                    student_info['department_code'] = self.department_code
                    
                    semesters = parse_semesters_with_offset(ws)
                    
                    student_data = {
                        "student": student_info,
                        "semesters": semesters,
                        "sheet_name": sheet_name,
                        "parsed_at": datetime.utcnow().isoformat()
                    }
                    
                    if student_data.get('student', {}).get('id'):
                        students.append(student_data)
                        self.stats["students"] += 1
                        self.stats["semesters"] += len(semesters)
                        for sem in semesters:
                            self.stats["courses"] += len(sem.get("courses", []))
                        
                        # Save to Supabase if available
                        if self._get_supabase():
                            self._save_to_supabase(student_data)
                    else:
                        errors.append({
                            "sheet": sheet_name,
                            "error": "No valid student ID found"
                        })
                        
                except Exception as e:
                    error_msg = str(e)
                    logger.error(f"Error parsing sheet {sheet_name}: {error_msg}")
                    errors.append({
                        "sheet": sheet_name,
                        "error": error_msg
                    })
                    continue
            
            wb.close()
            logger.info(f"Parsing complete: {self.stats}")
            
        except Exception as e:
            logger.error(f"Failed to open workbook: {e}")
            errors.append({
                "sheet": "all",
                "error": f"Failed to open workbook: {str(e)}"
            })
            
        return students, errors
    
    def _save_to_supabase(self, student_data: Dict):
        """Save student data to Supabase with all fields populated"""
        supabase = self._get_supabase()
        if not supabase:
            return
        
        student_info = student_data["student"]
        student_code = student_info.get("id")
        semesters = student_data.get("semesters", [])
        
        if not student_code:
            logger.warning(f"Skipping student without ID: {student_info.get('name')}")
            return
        
        try:
            # 1. استخراج سنة الالتحاق من أول ترم
            enrollment_year = None
            cumulative_gpa = None
            study_level_num = None
            
            if semesters:
                # سنة الالتحاق من أول ترم
                first_semester = semesters[0]
                academic_year = first_semester.get("academic_year", "")
                if "-" in academic_year:
                    try:
                        enrollment_year = int(academic_year.split("-")[0])
                    except:
                        pass
                
                # GPA من آخر ترم (هو التراكمي)
                last_semester = semesters[-1]
                if last_semester.get("gpa"):
                    try:
                        cumulative_gpa = float(last_semester.get("gpa", 0))
                    except:
                        pass
            
            # 2. استخراج رقم المستوى من النص
            level_str = student_info.get("study_level", "")
            level_map = {"الأول": 1, "الثاني": 2, "الثالث": 3, "الرابع": 4}
            for key, value in level_map.items():
                if key in level_str:
                    study_level_num = value
                    break
            
            # 3. الحصول على department_id من الكود
            department_id = self._get_department_id()
            
            # 4. إدراج الطالب بكل البيانات
            student_result = supabase.table("students").insert({
                "student_code": student_code,
                "name": student_info.get("name", ""),
                "department_id": department_id,
                "enrollment_year": enrollment_year,
                "study_level": study_level_num,
                "cumulative_gpa": cumulative_gpa,
                "cumulative_percentage": self._parse_percentage(student_info.get("cumulative_percentage", "")),
                "is_active": True,
            }).execute()
            
            student_db_id = student_result.data[0]["id"]
            logger.info(f"✅ Inserted student {student_code} (level: {study_level_num}, year: {enrollment_year}, GPA: {cumulative_gpa})")
            
            # 5. إدراج الفصول والمواد
            for idx, semester in enumerate(student_data.get("semesters", [])):
                semester_result = supabase.table("student_semesters").insert({
                    "student_id": student_db_id,
                    "semester_number": idx + 1,
                    "academic_year": semester.get("academic_year", ""),
                    "total_hours": self._parse_int(semester.get("semester_hours")),
                    "semester_gpa": self._parse_float(semester.get("semester_gpa")),
                }).execute()
                
                semester_db_id = semester_result.data[0]["id"]
                
                for course in semester.get("courses", []):
                    grade_points = self._calculate_grade_points(course.get("grade_letter", ""))
                    passed = course.get("passed", "نعم") in ["نعم", "√", "ناجح", True]
                    
                    supabase.table("student_courses").insert({
                        "student_id": student_db_id,
                        "semester_id": semester_db_id,
                        "course_code": course.get("course_code", ""),
                        "course_name": course.get("course_name", ""),
                        "credit_hours": self._parse_int(course.get("hours")),
                        "credit_hours_attempted": self._parse_int(course.get("hours")),
                        "passed": passed,
                        "grade_letter": course.get("grade_letter", ""),
                        "grade_points": grade_points,
                        "score": self._parse_float(course.get("score")),
                    }).execute()
                    
            logger.debug(f"Successfully saved student {student_code} with {len(semesters)} semesters")
            
        except Exception as e:
            logger.error(f"Failed to save student {student_code} to Supabase: {e}")
            self.errors.append({
                "student_code": student_code,
                "error": str(e)
            })
    
    def _parse_int(self, value) -> int:
        """Safely parse integer from string"""
        if not value:
            return 0
        try:
            return int(float(str(value)))
        except:
            return 0
    
    def _parse_float(self, value) -> float:
        """Safely parse float from string"""
        if not value:
            return None
        try:
            return float(str(value))
        except:
            return None
    
    def _parse_percentage(self, value) -> float:
        """Parse percentage string like '72.5%' to float"""
        if not value:
            return None
        try:
            return float(str(value).replace("%", "").strip())
        except:
            return None
    
    def _calculate_grade_points(self, grade_letter: str) -> float:
        """Convert grade letter to grade points"""
        grade_map = {
            "A+": 4.0, "A": 4.0, "A-": 3.7,
            "B+": 3.3, "B": 3.0, "B-": 2.7,
            "C+": 2.3, "C": 2.0, "C-": 1.7,
            "D+": 1.3, "D": 1.0, "D-": 0.7,
            "F": 0.0
        }
        return grade_map.get(grade_letter, 0.0)
    
    def get_stats(self) -> Dict:
        """Get parsing statistics"""
        return self.stats