# ===== File: core/security/validators.py (معدل) =====

from pathlib import Path
import openpyxl

class ExcelValidator:
    """Excel file validation"""
    
    REQUIRED_COLUMNS = ['CD', 'BC', 'BU', 'AL', 'O', 'Q', 'I']  # Reference only
    
    @staticmethod
    def validate_file(file_path: Path) -> tuple[bool, list[str]]:
        """
        Validate file format and structure
        Returns: (is_valid, errors_list)
        """
        errors = []
        
        # 1. Check file extension
        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            errors.append(f"Invalid file extension: {file_path.suffix}")
            return False, errors
        
        # 2. Check file size
        file_size = file_path.stat().st_size
        max_size = 50 * 1024 * 1024  # 50MB
        if file_size > max_size:
            errors.append(f"File too large: {file_size / 1024 / 1024:.2f}MB > 50MB")
            return False, errors
        
        # 3. Try to open and validate basic structure
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            
            if not sheets:
                errors.append("No sheets found in workbook")
                return False, errors
            
            # Check at least one sheet has data
            has_data = False
            for sheet_name in sheets[:3]:  # Check first 3 sheets
                ws = wb[sheet_name]
                # Check if there's any non-empty cell
                for row in ws.iter_rows(max_row=10, max_col=10, values_only=True):
                    if any(cell for cell in row):
                        has_data = True
                        break
                if has_data:
                    break
            
            if not has_data:
                errors.append("No data found in workbook")
            
            wb.close()
            
        except Exception as e:
            errors.append(f"Cannot read Excel file: {str(e)}")
            return False, errors
        
        return True, errors